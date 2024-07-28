"""
This module provides functionality for interacting with Excel files.

It includes functions to:
- Read data from Excel files.
- Write data to Excel files.
- Update existing Excel files.
- Perform various data manipulations and transformations.

Functions:
- read_excel: Reads data from an Excel file and returns it as a DataFrame.
- write_excel: Writes data from a DataFrame to an Excel file.
- update_excel: Updates specific cells or ranges in an existing Excel file.
- transform_data: Performs data transformations on a DataFrame.
- validate_excel: Validates the structure and content of an Excel file.
"""

import xlsxwriter
import openpyxl

from app.config import FILE_EXCEL
from app.cell_formats import header_format_dict
from app.cell_formats import wallets_column_format_dict
from app.cell_formats import total_cell_format_dict
from app.cell_formats import common_ceil_format_dict
from app.cell_formats import usd_ceil_format_dict
from app.cell_formats import donate_cell_format_dict

def adjust_column_width(filename):
    """
    Adjusts the width of columns in an Excel worksheet based on the maximum length of the content in each column.

    Args:
        filename (str): The path to the Excel file to be adjusted.

    Returns:
        None
    """
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    column_max_length = [0] * sheet.max_column

    ya_ustal = 0
    for row in sheet.iter_rows(values_only=True):
        if ya_ustal != 0:
            for idx, cell_value in enumerate(row):
                width = 0
                if cell_value is not None:
                    strs = str(cell_value).split('\n')
                    for i in strs:
                        width = max(width, len(i))

                    cell_length = width
                    if cell_length > column_max_length[idx]:
                        column_max_length[idx] = cell_length
        ya_ustal += 1
    for col_idx, max_length in enumerate(column_max_length[1:], start=2):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = 10 if max_length < 10 else max_length

    workbook.save(filename)


def save_full_to_excel(wallets, chains, coins, balances):
    """
    Saves wallet, chain, coin, and balance data to an Excel file.

    Args:
        wallets (list): A list of wallet data to be saved.
        chains (list): A list of chain data to be saved.
        coins (list): A list of coin data to be saved.
        balances (list): A list of balance data to be saved.

    Returns:
        None
    """
    workbook = xlsxwriter.Workbook(FILE_EXCEL)
    worksheet = workbook.add_worksheet("Coins")

    header_format = workbook.add_format(header_format_dict)
    wallets_column_format = workbook.add_format(wallets_column_format_dict)
    total_cell_format = workbook.add_format(total_cell_format_dict)
    common_ceil_format = workbook.add_format(common_ceil_format_dict)
    usd_ceil_format = workbook.add_format(usd_ceil_format_dict)
    donate_cell_format = workbook.add_format(donate_cell_format_dict)

    headers = ['Wallet', *[chain.upper() for chain in chains], 'CHAINS', 'TOTAL']

    for row_id, wallet in enumerate(wallets):
        worksheet.write(row_id + 1, 0, wallet, wallets_column_format)
    worksheet.write(len(wallets) + 1, 0, 'TOTAL IN USD', total_cell_format)


    for col_id, chain in enumerate(headers):
        worksheet.write(0, col_id, chain, header_format)


    for col_id, chain in enumerate(chains):
        total_in_chain = 0.0
        for row_id, wallet in enumerate(wallets):
            cell = ''
            for coin in coins[chain][wallet]:
                coin_in_usd = '?' if (coin["price"] is None) else round(coin["amount"] * coin["price"], 2)
                cell += f'{coin["ticker"]} - {round(coin["amount"], 4)} (${coin_in_usd})\n'
                total_in_chain += coin_in_usd if isinstance(coin_in_usd, float) else 0
            if cell == '':
                cell = '--'
            cell = cell[:-1]
            worksheet.write(row_id + 1, col_id + 1, cell, common_ceil_format)
        worksheet.write(len(wallets) + 1, col_id + 1, f'${round(total_in_chain, 2)}', usd_ceil_format)


    total_usd = 0.0
    total_all_chains = 0.0
    for row_id, wallet in enumerate(wallets):
        total_in_wallet = 0.0
        for chain in chains:
            for coin in coins[chain][wallet]:
                coin_in_usd = 0 if (coin["price"] is None) else round(coin["amount"] * coin["price"], 2)
                total_in_wallet += coin_in_usd
        total_usd += total_in_wallet
        total_all_chains += balances[wallet]
        worksheet.write(row_id + 1, len(headers) - 2, f'${round(total_in_wallet, 2)}', usd_ceil_format)
        worksheet.write(row_id + 1, len(headers) - 1, f'${round(balances[wallet], 2)}', usd_ceil_format)
    worksheet.write(len(wallets) + 1, len(headers) - 2, f'${round(total_usd, 2)}', usd_ceil_format)
    worksheet.write(len(wallets) + 1, len(headers) - 1, f'${round(total_all_chains, 2)}', usd_ceil_format)


    worksheet.write(len(wallets) + 3, 0, 'Donate:', donate_cell_format)
    worksheet.write(len(wallets) + 4, 0, '0x2e69Da32b0F7e75549F920CD2aCB0532Cc2aF0E7', donate_cell_format)


    worksheet.set_row(0, 35)
    worksheet.set_column(0, 0, 52)

    workbook.close()

    adjust_column_width(FILE_EXCEL)


def save_selected_to_excel(wallets, chains, coins, balances, ticker):
    """
    Create an Excel file with selected cryptocurrency data.

    Args:
    wallets (list): List of wallet addresses.
    chains (list): List of blockchain networks.
    coins (dict): Nested dictionary of coin data for each wallet and chain.
    balances (dict): Dictionary of total balances for each wallet.
    ticker (str): The ticker symbol of the cryptocurrency to focus on.

    This function orchestrates the creation of the Excel file by calling
    helper functions for different parts of the spreadsheet.
    """
    workbook = xlsxwriter.Workbook(FILE_EXCEL)
    worksheet = workbook.add_worksheet("Coins")

    formats = {
        'header': workbook.add_format(header_format_dict),
        'wallets_column': workbook.add_format(wallets_column_format_dict),
        'total_cell': workbook.add_format(total_cell_format_dict),
        'common_cell': workbook.add_format(common_ceil_format_dict),
        'usd_cell': workbook.add_format(usd_ceil_format_dict),
        'donate_cell': workbook.add_format(donate_cell_format_dict)
    }

    headers = ['Wallet'] + [chain.upper() for chain in chains] + ['CHAINS', 'TOTAL']

    write_headers(worksheet, headers, formats['header'])
    write_wallets(worksheet, wallets, formats['wallets_column'], formats['total_cell'])
    
    total_usd, total_all_chains = write_data(worksheet, wallets, chains, coins, balances, ticker, formats)
    
    write_totals(worksheet, len(wallets), len(headers), total_usd, total_all_chains, formats['usd_cell'])
    write_donation_info(worksheet, len(wallets), formats['donate_cell'])

    format_worksheet(worksheet)
    workbook.close()
    adjust_column_width(FILE_EXCEL)

def write_headers(worksheet, headers, header_format):
    """
    Write the header row to the worksheet.

    Args:
    worksheet (Worksheet): The Excel worksheet object.
    headers (list): List of header titles.
    header_format (Format): The Excel cell format for headers.

    This function writes the headers and merges cells for multi-column headers.
    """
    for col_id, header in enumerate(headers):
        if col_id == 0 or col_id >= len(headers) - 2:
            worksheet.write(0, col_id + (len(headers) - 3) * 2, header, header_format)
        else:
            worksheet.merge_range(0, col_id - 2 + 2 * col_id, 0, col_id + 2 * col_id, header, header_format)

def write_wallets(worksheet, wallets, wallet_format, total_format):
    """
    Write the wallet addresses and the 'TOTAL IN USD' row.

    Args:
    worksheet (Worksheet): The Excel worksheet object.
    wallets (list): List of wallet addresses.
    wallet_format (Format): The Excel cell format for wallet addresses.
    total_format (Format): The Excel cell format for the total row.
    """
    for row_id, wallet in enumerate(wallets, start=1):
        worksheet.write(row_id, 0, wallet, wallet_format)
    worksheet.write(len(wallets) + 1, 0, 'TOTAL IN USD', total_format)

def write_data(worksheet, wallets, chains, coins, balances, ticker, formats):
    """
    Write the main data to the worksheet.

    Args:
    worksheet (Worksheet): The Excel worksheet object.
    wallets (list): List of wallet addresses.
    chains (list): List of blockchain networks.
    coins (dict): Nested dictionary of coin data for each wallet and chain.
    balances (dict): Dictionary of total balances for each wallet.
    ticker (str): The ticker symbol of the cryptocurrency to focus on.
    formats (dict): Dictionary of Excel cell formats.

    Returns:
    tuple: Total USD value and total balance across all chains.

    This function writes the coin data for each wallet and chain, and calculates totals.
    """
    total_usd = 0.0
    total_all_chains = 0.0

    for col_id, chain in enumerate(chains):
        total_in_chain, total_amount = write_chain_data(worksheet, wallets, chain, coins, ticker, col_id, formats)
        write_chain_totals(worksheet, len(wallets), col_id, ticker, total_amount, total_in_chain, formats['usd_cell'])

    for row_id, wallet in enumerate(wallets, start=1):
        total_in_wallet = sum(coin["amount"] * (coin["price"] or 0) 
                              for chain in chains 
                              for coin in coins[chain][wallet] 
                              if coin['ticker'] == ticker)
        total_usd += total_in_wallet
        total_all_chains += balances[wallet]
        write_wallet_totals(worksheet, row_id, len(chains), total_in_wallet, balances[wallet], formats['usd_cell'])

    return total_usd, total_all_chains

def write_chain_data(worksheet, wallets, chain, coins, ticker, col_id, formats):
    """
    Write data for a specific blockchain.

    Args:
    worksheet (Worksheet): The Excel worksheet object.
    wallets (list): List of wallet addresses.
    chain (str): The current blockchain being processed.
    coins (dict): Nested dictionary of coin data for each wallet and chain.
    ticker (str): The ticker symbol of the cryptocurrency to focus on.
    col_id (int): The column index for the current chain.
    formats (dict): Dictionary of Excel cell formats.

    Returns:
    tuple: Total value in the chain and total amount of the coin.

    This function writes coin data for each wallet in a specific blockchain.
    """
    total_in_chain = 0.0
    total_amount = 0.0
    for row_id, wallet in enumerate(wallets, start=1):
        coin_data = next((coin for coin in coins[chain][wallet] if coin['ticker'] == ticker), None)
        if coin_data:
            amount = coin_data['amount']
            coin_in_usd = '?' if coin_data["price"] is None else round(amount * coin_data["price"], 2)
            total_in_chain += coin_in_usd if isinstance(coin_in_usd, float) else 0
            total_amount += amount
        else:
            amount, coin_in_usd = 0, 0
        write_coin_data(worksheet, row_id, col_id, ticker, amount, coin_in_usd, formats['common_cell'])
    return total_in_chain, total_amount

def write_coin_data(worksheet, row, col, ticker, amount, coin_in_usd, cell_format):
    """
    Write coin data for a single cell.

    Args:
    worksheet (Worksheet): The Excel worksheet object.
    row (int): The row index.
    col (int): The column index.
    ticker (str): The ticker symbol of the cryptocurrency.
    amount (float): The amount of the coin.
    coin_in_usd (float or str): The USD value of the coin amount.
    cell_format (Format): The Excel cell format to use.

    This function writes the ticker, amount, and USD value for a single coin entry.
    """
    base_col = col - 1 + (col + 1) * 2
    worksheet.write(row, base_col, ticker, cell_format)
    worksheet.write(row, base_col + 1, round(amount, 4), cell_format)
    worksheet.write(row, base_col + 2, f'${coin_in_usd}', cell_format)

def write_chain_totals(worksheet, num_wallets, col, ticker, total_amount, total_in_chain, cell_format):
    """
    Write totals for a specific blockchain.

    Args:
    worksheet (Worksheet): The Excel worksheet object.
    num_wallets (int): The number of wallets.
    col (int): The column index for the current chain.
    ticker (str): The ticker symbol of the cryptocurrency.
    total_amount (float): The total amount of the coin in this chain.
    total_in_chain (float): The total USD value in this chain.
    cell_format (Format): The Excel cell format to use.

    This function writes the total amount and value for a specific blockchain.
    """
    base_col = col - 1 + (col + 1) * 2
    worksheet.write(num_wallets + 1, base_col, ticker, cell_format)
    worksheet.write(num_wallets + 1, base_col + 1, round(total_amount, 4), cell_format)
    worksheet.write(num_wallets + 1, base_col + 2, f'${round(total_in_chain, 2)}', cell_format)

def write_wallet_totals(worksheet, row, num_chains, total_in_wallet, total_balance, cell_format):
    """
    Write totals for a specific wallet.

    Args:
    worksheet (Worksheet): The Excel worksheet object.
    row (int): The row index for the current wallet.
    num_chains (int): The number of blockchains.
    total_in_wallet (float): The total USD value in this wallet for the specific coin.
    total_balance (float): The total balance of the wallet across all coins.
    cell_format (Format): The Excel cell format to use.

    This function writes the total value and balance for a specific wallet.
    """
    base_col = num_chains * 3
    worksheet.write(row, base_col, f'${round(total_in_wallet, 2)}', cell_format)
    worksheet.write(row, base_col + 1, f'${round(total_balance, 2)}', cell_format)

def write_totals(worksheet, num_wallets, num_headers, total_usd, total_all_chains, cell_format):
    """
    Write the grand totals at the bottom of the sheet.

    Args:
    worksheet (Worksheet): The Excel worksheet object.
    num_wallets (int): The number of wallets.
    num_headers (int): The number of header columns.
    total_usd (float): The total USD value across all wallets for the specific coin.
    total_all_chains (float): The total balance across all wallets and all coins.
    cell_format (Format): The Excel cell format to use.

    This function writes the grand total USD value and balance at the bottom of the sheet.
    """
    base_col = (num_headers - 3) * 2
    worksheet.write(num_wallets + 1, base_col, f'${round(total_usd, 2)}', cell_format)
    worksheet.write(num_wallets + 1, base_col + 1, f'${round(total_all_chains, 2)}', cell_format)

def write_donation_info(worksheet, num_wallets, cell_format):
    """
    Write donation information at the bottom of the sheet.

    Args:
    worksheet (Worksheet): The Excel worksheet object.
    num_wallets (int): The number of wallets.
    cell_format (Format): The Excel cell format to use for donation info.

    This function adds donation information at the bottom of the spreadsheet.
    """
    worksheet.write(num_wallets + 3, 0, 'Donate:', cell_format)
    worksheet.write(num_wallets + 4, 0, '0x2e69Da32b0F7e75549F920CD2aCB0532Cc2aF0E7', cell_format)

def format_worksheet(worksheet):
    """
    Apply final formatting to the worksheet.

    Args:
    worksheet (Worksheet): The Excel worksheet object.

    This function sets the row height for the header and adjusts the width of the first column.
    """
    worksheet.set_row(0, 35)
    worksheet.set_column(0, 0, 52)
